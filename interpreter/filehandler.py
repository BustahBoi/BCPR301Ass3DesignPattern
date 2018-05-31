from pathlib import Path, PurePosixPath
from abc import ABCMeta, abstractmethod
from csv import DictReader as CSVDictReader
from openpyxl import load_workbook, utils
from datetime import datetime, date
from interpreter.validator import Validator
from interpreter import validator


# James
class FileHandler:
    # James
    def __init__(self):
        self.filename = None
        self.file_type = None

    def load(self, filename):
        self.filename = filename
        self.set_file_type()

    # James
    def file_exist(self):
        """
        Checks if the current file exists
        """
        if self.filename.exists():
            return True
        else:
            return False

    # Wesley
    def set_file_type(self):
        """
        Will get the file type and will create the
        corresponding solid class and set it to self.file_type
        """
        suffix = PurePosixPath(self.filename).suffix
        print(suffix)
        file_types = {
            '.csv': FileTypeCSV(),
            '.xlsx': FileTypeXLSX(),
            '.txt': FileTypeTXT()
        }
        self.file_type = file_types[suffix]

    def read(self):
        return self.file_type.read(self.filename)


# Wesley
class FileTypeAbstract(metaclass=ABCMeta):
    # Wesley
    @abstractmethod
    def read(self, filename):
        pass


# Wesley
class FileTypeCSV(FileTypeAbstract):
    # James
    def read(self, filename):
        """
        Return dictionary with key => value pairs
        :param filename is the file where the values exist
        >>> read("Saves/data.csv")
        """
        # try:
        data = dict()
        empno = 0
        with open(filename) as f:
            reader = CSVDictReader(f)
            for row in reader:
                record = dict()
                for key in row:
                    record[key] = row.get(key)
                data[empno] = record
                empno += 1
            # print(data)
        # James' changes (13/03)
        result = Validator.save_dict(data)
        print(result)
        return result
        # except TypeError:
        #     print("Error!!")


# Wesley
class FileTypeXLSX(FileTypeAbstract):
    # Wesley
    def read(self, filename):
        """
        Return dictionary with key => value pairs
        :param filename is the file where the values exist
        >>> f = FileTypeXLSX()
        >>> f.read("Saves/doctest.xlsx")
        {1: {'ID': 'G262', 'Gender': 'Female', 'Age': 12, 'Sales': 215, 'BMI': 'Normal', 'Salary': 23, 'Birthday': '24-05-1993'}, 2: {'ID': 'A233', 'Gender': 'Chihuahua', 'Age': 22, 'Sales': 245, 'BMI': 'normal', 'Salary': 23, 'Birthday': '24-06-1995'}, 3: {'ID': 'A262', 'Gender': 'M', 'Age': 24, 'Sales': 845, 'BMI': 'Normal', 'Salary': 23, 'Birthday': '24-05-1993'}, 4: {'ID': 'A233', 'Gender': 'Female', 'Age': 62, 'Sales': 245, 'BMI': 'normal', 'Salary': 23, 'Birthday': '24-06-1995'}}
        Adding Row 1
        {'ID': 'G262', 'Gender': 'F', 'Age': '12', 'Sales': '215', 'BMI': 'Normal', 'Salary': '23', 'Birthday': '24-05-1993'}
        Error at entry: 2
        Adding Row 3
        {'ID': 'A262', 'Gender': 'M', 'Age': '24', 'Sales': '845', 'BMI': 'Normal', 'Salary': '23', 'Birthday': '24-05-1993'}
        Adding Row 4
        {'ID': 'A233', 'Gender': 'F', 'Age': '62', 'Sales': '245', 'BMI': 'Normal', 'Salary': '23', 'Birthday': '24-06-1995'}
        result
        {1: {'ID': 'G262', 'Gender': 'F', 'Age': '12', 'Sales': '215', 'BMI': 'Normal', 'Salary': '23', 'Birthday': '24-05-1993'}, 3: {'ID': 'A262', 'Gender': 'M', 'Age': '24', 'Sales': '845', 'BMI': 'Normal', 'Salary': '23', 'Birthday': '24-05-1993'}, 4: {'ID': 'A233', 'Gender': 'F', 'Age': '62', 'Sales': '245', 'BMI': 'Normal', 'Salary': '23', 'Birthday': '24-06-1995'}}
            """
        data = dict()
        empno = 0
        keys = []
        a_row = 0
        try:
            workbook = load_workbook(filename)
            first_sheet = workbook.sheetnames[0]
            worksheet = workbook[first_sheet]
            for row in worksheet.iter_rows():
                record = dict()
                row_num = 0
                for cell in row:
                    a_row = cell.row
                    if 1 == a_row:
                        keys.append(cell.value)
                    else:
                        valid = cell.value
                        if isinstance(cell.value, datetime):
                            valid = Validator.xlsx_date(cell.value)
                        record[keys[row_num]] = valid
                    row_num += 1
                if a_row > 1:
                    data[empno] = record
                empno += 1
            # print(data)
            result = Validator.save_dict(data)
            validator.a = None
        except PermissionError:
            result = dict()
            print("Sorry, you don't have enough permissions to access this file")
        print("result")
        return result
# The above function contains a date object in the dictionary for each date,
# as the birthday is a date, may need to access the values stored in the date object when validating


# Sam
class FileTypeTXT(FileTypeAbstract):
    def read(self, filename):
        """
        Return dictionary with key => value pairs
        :param filename is the file where the values exist
        >>> read("Saves/data.txt")
        """
        empno = 0
        try:
            file = open(filename, 'r')
            print(file)
            for line in file:
                print(line)
                dictionary = dict()
                rows = line.split(":")
                for row in rows:
                    if len(row.split("=")) == 2:
                        key = row.split("=")[0]
                        value = row.split("=")[1]
                        value = value.rstrip('\n')
                        dictionary[key] = value
                    else:
                        print("File error")
                        return False
            return dictionary
        finally:
            print("something weird happened... guys pls send help")


# def run():
#     a = FileHandler.get_file_name()
#     aclass = FileHandler(a)
#     while aclass.file_exist() is False:
#         print("File exists:", aclass.file_exist())
#         a = FileHandler.get_file_name()
#         aclass = FileHandler(a)
#     aclass.set_file_type()
#     aclass.read()


# run()